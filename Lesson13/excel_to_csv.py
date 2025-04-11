import sys
import csv
from openpyxl import load_workbook
from io import BytesIO
import chardet

# Чтение Excel-файла из STDIN (двоичные данные)
input_data = sys.stdin.buffer.read()
excel_file = BytesIO(input_data)

# Загрузка Excel-файла
wb = load_workbook(excel_file, read_only=True, data_only=True)
ws = wb.active  # Используем первый лист

# Преобразование данных в CSV
output = sys.stdout  # Используем текстовый поток
writer = csv.writer(output, lineterminator='\n')  # Устанавливаем Unix-перенос строк

for row in ws.iter_rows(values_only=True):
    encoded_row = []
    for cell in row:
        if isinstance(cell, str):
            # Анализируем байты строки напрямую
            try:
                detected = chardet.detect(cell.encode('utf-8', errors='ignore'))
                encoding = detected.get('encoding', 'utf-8')  # По умолчанию utf-8

                # Преобразуем строку в utf-8
                encoded_cell = cell.encode(encoding, errors='ignore').decode('utf-8')
            except (UnicodeEncodeError, UnicodeDecodeError):
                # Если возникает ошибка, оставляем строку как есть
                encoded_cell = cell
            encoded_row.append(encoded_cell)
        else:
            # Для чисел или пустых значений используем строковое представление
            encoded_row.append(str(cell or ''))
    writer.writerow(encoded_row)

# Закрываем файл
wb.close()